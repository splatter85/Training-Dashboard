import streamlit as st
import pandas as pd
from io import BytesIO
import datetime
import base64
import requests
import json
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font

st.set_page_config(page_title="Training Dashboard Uploader", layout="centered")
st.title("📊 Training Dashboard Uploader")

# --- GitHub Config ---
GITHUB_REPO = "splatter85/Training-Dashboard"
GITHUB_BRANCH = "main"
DASHBOARD_FILE_PATH = "Training_Dashboard.csv"

# Column colors for styling
COLUMN_COLORS = {
    'Total Trainings': "#e1f7d5",
    'DXFleet': "#ffbdbd",
    'Phoenix SQL Lite': "#c9c9ff",
    'Cancellations': "#f1cbff",
    'No-Shows': "#f7c6c7",
    'Pacific': "#aec6cf",
    'Mountain': "#77dd77",
    'Central': "#ffd1dc",
    'Eastern': "#ffffb3"
}


def get_initial_dashboard():
    try:
        url = f"https://raw.githubusercontent.com/{GITHUB_REPO}/{GITHUB_BRANCH}/{DASHBOARD_FILE_PATH}"
        return pd.read_csv(url)
    except:
        return pd.DataFrame(columns=['Month', 'Total Trainings', 'DXFleet', 'Phoenix SQL Lite', 'Cancellations', 'No-Shows', 'Pacific', 'Mountain', 'Central', 'Eastern'])


def push_updated_dashboard_to_github(updated_df):
    token = st.secrets["GITHUB_TOKEN"]
    headers = {
        "Authorization": f"token {token}",
        "Accept": "application/vnd.github.v3+json"
    }

    api_url = f"https://api.github.com/repos/{GITHUB_REPO}/contents/{DASHBOARD_FILE_PATH}"
    get_resp = requests.get(api_url, headers=headers)
    sha = get_resp.json().get("sha") if get_resp.status_code == 200 else None

    csv_data = updated_df.to_csv(index=False)
    encoded_content = base64.b64encode(csv_data.encode()).decode()

    payload = {
        "message": "Update Training Dashboard via Streamlit",
        "content": encoded_content,
        "branch": GITHUB_BRANCH
    }
    if sha:
        payload["sha"] = sha

    put_resp = requests.put(api_url, headers=headers, data=json.dumps(payload))
    if put_resp.status_code in [200, 201]:
        st.success("✅ Dashboard saved to GitHub!")
    else:
        st.error(f"❌ Failed to save dashboard to GitHub: {put_resp.json()}")


def style_excel_file(df):
    wb = Workbook()
    ws = wb.active
    ws.title = "Dashboard"

    # Write headers
    for col_num, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_num, value=col_name)
        cell.font = Font(bold=True)

    # Write data and apply column coloring
    for row_idx, row in df.iterrows():
        for col_idx, (col_name, value) in enumerate(row.items(), 1):
            cell = ws.cell(row=row_idx + 2, column=col_idx, value=value)
            if str(value).strip().upper() == "TOTAL":
                for style_cell in ws[row_idx + 2]:
                    style_cell.font = Font(bold=True)
                    style_cell.fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
            else:
                fill_color = COLUMN_COLORS.get(col_name, None)
                if fill_color:
                    cell.fill = PatternFill(start_color=fill_color[1:], end_color=fill_color[1:], fill_type="solid")

    stream = BytesIO()
    wb.save(stream)
    return stream.getvalue()

# Load dashboard
dashboard_df = get_initial_dashboard()

# Remove total rows if they exist
if not dashboard_df.empty and dashboard_df.iloc[-1][0] == 'TOTAL':
    dashboard_data = dashboard_df.iloc[:-2]
else:
    dashboard_data = dashboard_df.copy()

# Add total row
if not dashboard_data.empty:
    numeric_cols = dashboard_data.columns.drop('Month')
    totals = dashboard_data[numeric_cols].sum(numeric_only=True)
    total_row = pd.DataFrame([[''] + totals.tolist()])
    total_row.columns = dashboard_data.columns
    labeled_total_row = total_row.copy()
    labeled_total_row.iloc[0, 0] = 'TOTAL'
    dashboard_display = pd.concat([dashboard_data, pd.DataFrame([[''] * len(dashboard_data.columns)], columns=dashboard_data.columns), labeled_total_row], ignore_index=True)
else:
    dashboard_display = dashboard_df

# Style table for on-screen display
styled_df = dashboard_display.style
for col, color in COLUMN_COLORS.items():
    if col in dashboard_display.columns:
        styled_df = styled_df.set_properties(subset=[col], **{'background-color': color})

# Format numbers as integers
styled_df = styled_df.format({col: '{:.0f}' for col in dashboard_display.columns if col != 'Month'})

st.subheader("Current Dashboard")
st.dataframe(styled_df, use_container_width=True)

# Downloads
st.download_button("Download Dashboard as CSV", data=dashboard_display.to_csv(index=False), file_name="Training_Dashboard.csv")
st.download_button("Download Dashboard as Excel", data=style_excel_file(dashboard_display), file_name="Training_Dashboard.xlsx")

st.markdown("---")
uploaded_file = st.file_uploader("Upload a monthly training report (CSV)", type=["csv"])

if uploaded_file:
    try:
        df = pd.read_csv(uploaded_file)
        df['Start Date & Time'] = pd.to_datetime(df['Start Date & Time'], errors='coerce')
        df['Month'] = df['Start Date & Time'].dt.to_period('M').astype(str)

        if df['Month'].nunique() != 1:
            st.error("The file must contain data from only one month.")
        else:
            target_month = df['Month'].iloc[0]
            if target_month in dashboard_data['Month'].values:
                st.error(f"Data from {target_month} is already in the dashboard.")
            else:
                num_dxfleet = df['Event Type Name'].str.contains('DXFleet', case=False, na=False).sum()
                num_phoenix = df['Event Type Name'].str.contains('Phoenix SQL Lite', case=False, na=False).sum()
                num_cancellations = (df['Canceled'] == True).sum() if 'Canceled' in df.columns else 0
                num_noshows = (df['Marked as No-Show'].str.lower() == 'yes').sum()
                tz_counts = df['Invitee Time Zone'].value_counts()
                tz_pacific = tz_counts.get('Pacific Time - US & Canada', 0)
                tz_mountain = tz_counts.get('Mountain Time - US & Canada', 0)
                tz_central = tz_counts.get('Central Time - US & Canada', 0)
                tz_eastern = tz_counts.get('Eastern Time - US & Canada', 0)

                new_row = {
                    'Month': target_month,
                    'Total Trainings': len(df),
                    'DXFleet': num_dxfleet,
                    'Phoenix SQL Lite': num_phoenix,
                    'Cancellations': num_cancellations,
                    'No-Shows': num_noshows,
                    'Pacific': tz_pacific,
                    'Mountain': tz_mountain,
                    'Central': tz_central,
                    'Eastern': tz_eastern
                }

                updated_data = pd.concat([dashboard_data, pd.DataFrame([new_row])], ignore_index=True)
                totals = updated_data.drop(columns='Month').sum(numeric_only=True)
                total_row = pd.DataFrame([[''] + totals.tolist()], columns=updated_data.columns)
                total_row.iloc[0, 0] = 'TOTAL'
                final_dashboard = pd.concat([updated_data, pd.DataFrame([[''] * len(updated_data.columns)], columns=updated_data.columns), total_row], ignore_index=True)

                push_updated_dashboard_to_github(final_dashboard)
                st.experimental_rerun()

    except Exception as e:
        st.error(f"An error occurred while processing the file: {e}")
